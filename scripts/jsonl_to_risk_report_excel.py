#!/usr/bin/env python3
"""
将线索 JSONL 转为风险报送批量上传 Excel（17 列，与 数据模版.xlsx / RiskReportExcelParser 一致）。

用法:
  python3 scripts/jsonl_to_risk_report_excel.py
  python3 scripts/jsonl_to_risk_report_excel.py --jsonl ai-file/xxx.jsonl --output ai-file/xxx.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_JSONL = REPO_ROOT / "ai-file" / "人工智能安全风险事件列表v2（汇总版）-至0521_1.jsonl"
DEFAULT_TEMPLATE = REPO_ROOT / "ai-file" / "数据模版.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "ai-file" / "人工智能安全风险事件列表v2（汇总版）-至0521_1_批量上传.xlsx"

ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# 与 RiskReportExcelParser 列序一致（0-based）
COLUMNS = [
    ("number", "序号"),
    ("event_name", "事件名"),
    ("content", "内容"),
    ("class_report_1", "一级分类"),
    ("class_report_2", "二级分类"),
    ("products_components_services", "产品/组件/服务"),
    ("operating_entity", "运营主体"),
    ("risk_description", "风险描述"),
    ("source_url", "来源url"),
    ("source_website", "来源网站"),
    ("paper_title", "论文名称"),
    ("research_team", "研究团队"),
    ("is_verify", "是否验证"),
    ("is_submit", "是否报送"),
    ("submission_channel", "报送渠道"),
    ("submission_time", "报送时间"),
    ("submit_user_name", "报送人/分中心"),
]

DATETIME_FMT = "%Y-%m-%d %H:%M:%S"


def text_or_none(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return ILLEGAL_XML_CHARS.sub("", s)


def parse_submission_time(item: dict) -> datetime | None:
    raw = item.get("submission_time") or item.get("Submission_time")
    s = text_or_none(raw)
    if not s:
        return None
    for fmt in (DATETIME_FMT, "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == "%Y-%m-%d":
                return dt.replace(hour=0, minute=0, second=0)
            return dt
        except ValueError:
            continue
    raise ValueError(f"无法解析报送时间: {s!r}")


def row_value(item: dict, key: str):
    if key == "submission_time":
        return parse_submission_time(item)
    if key == "number":
        n = item.get("number")
        return int(n) if n is not None else None
    return text_or_none(item.get(key))


def load_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8") as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSONL 第 {lineno} 行解析失败: {exc}") from exc
    return rows


def verify_template_headers(ws) -> None:
    expected = [label for _, label in COLUMNS]
    actual = [ws.cell(1, c).value for c in range(1, len(expected) + 1)]
    if actual != expected:
        raise ValueError(
            "模板表头与解析器不一致\n"
            f"  期望: {expected}\n"
            f"  实际: {actual}"
        )


def convert(jsonl_path: Path, template_path: Path, output_path: Path) -> int:
    items = load_jsonl(jsonl_path)
    if not items:
        raise ValueError("JSONL 为空")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    verify_template_headers(ws)

    # 清空模板示例行
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    keys = [key for key, _ in COLUMNS]
    for i, item in enumerate(items, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(i, col_idx, value=row_value(item, key))

    wb.save(output_path)
    return len(items)


def main() -> int:
    parser = argparse.ArgumentParser(description="JSONL 转风险报送 Excel")
    parser.add_argument("--jsonl", type=Path, default=DEFAULT_JSONL)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.jsonl.is_file():
        raise SystemExit(f"找不到 JSONL: {args.jsonl}")
    if not args.template.is_file():
        raise SystemExit(f"找不到模板: {args.template}")

    count = convert(args.jsonl, args.template, args.output)
    print(f"已生成: {args.output}")
    print(f"共写入 {count} 条（17 列，表头与 数据模版.xlsx 一致）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
